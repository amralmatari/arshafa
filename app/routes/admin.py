"""
Admin routes for system administration
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, jsonify
from flask_login import login_required, current_user
from app import db
from app.models.user import User, Role, Permission
from app.models.document import Category, Document
from app.models.audit import AuditLog
import os

def translate_action(action):
    """Translate action names to Arabic"""
    translations = {
        'LOGIN': 'تسجيل دخول',
        'LOGOUT': 'تسجيل خروج',
        'LOGIN_FAILED': 'فشل تسجيل الدخول',
        'DOCUMENT_CREATE': 'إنشاء وثيقة',
        'DOCUMENT_UPDATE': 'تحديث وثيقة',
        'DOCUMENT_DELETE': 'حذف وثيقة',
        'DOCUMENT_VIEW': 'عرض وثيقة',
        'DOCUMENT_DOWNLOAD': 'تحميل وثيقة',
        'USER_CREATE': 'إنشاء مستخدم',
        'USER_UPDATE': 'تحديث مستخدم',
        'USER_DELETE': 'حذف مستخدم',
        'USER_ACTIVATE': 'تفعيل مستخدم',
        'USER_DEACTIVATE': 'إلغاء تفعيل مستخدم',
        'USER_ROLE_CHANGE': 'تغيير دور مستخدم',
        'CATEGORY_CREATE': 'إنشاء فئة',
        'CATEGORY_UPDATE': 'تحديث فئة',
        'CATEGORY_DELETE': 'حذف فئة',
        'ROLE_CREATE': 'إنشاء دور',
        'ROLE_UPDATE': 'تحديث دور',
        'ROLE_DELETE': 'حذف دور',
        'PERMISSION_UPDATE': 'تحديث صلاحية',
        'SYSTEM_BACKUP': 'نسخ احتياطي',
        'SYSTEM_RESTORE': 'استعادة نسخة احتياطية',
        'SETTINGS_UPDATE': 'تحديث الإعدادات'
    }
    return translations.get(action, action or 'غير محدد')
from app.utils.decorators import admin_required
from datetime import datetime
from functools import wraps
import json

bp = Blueprint('admin', __name__)

# Helper functions for category operations
def get_category_level(category):
    """Get category level in hierarchy"""
    level = 0
    current = category.parent
    while current:
        level += 1
        current = current.parent
    return level

def get_category_document_count(category_id):
    """Get total document count for a category including all subcategories recursively"""
    from sqlalchemy import func

    direct_count = db.session.query(func.count(Document.id))\
        .filter(Document.category_id == category_id)\
        .scalar() or 0

    subcategories = Category.query.filter_by(parent_id=category_id).all()
    recursive_count = 0
    for subcat in subcategories:
        recursive_count += get_category_document_count(subcat.id)

    return direct_count + recursive_count

def super_admin_required(f):
    """Decorator to require super admin privileges"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
            return redirect(url_for('main.index'))

        if not current_user.role or current_user.role.name != 'super_admin':
            flash('هذه الصفحة مخصصة لمدير النظام الرئيسي فقط', 'danger')
            return redirect(url_for('main.index'))

        return f(*args, **kwargs)
    return decorated_function

@bp.route('/')
@login_required
@admin_required
def index():
    """Admin dashboard"""
    # Get comprehensive statistics for the dashboard
    user_count = User.query.count()
    active_users = User.query.filter_by(is_active=True).count()
    category_count = Category.query.count()

    # Get document count if Document model exists
    try:
        from app.models.document import Document
        document_count = Document.query.count()
    except ImportError:
        document_count = 0

    # Get role count
    role_count = Role.query.count()

    # Get permission count
    permission_count = Permission.query.count()

    # Get admin users count
    admin_users = User.query.filter_by(is_admin=True).count()

    # Get recent activities count (today)
    from datetime import datetime, timedelta
    today = datetime.now().date()
    recent_activities = AuditLog.query.filter(
        AuditLog.created_at >= today
    ).count()

    return render_template('admin/index.html',
                          user_count=user_count,
                          active_users=active_users,
                          category_count=category_count,
                          document_count=document_count,
                          role_count=role_count,
                          permission_count=permission_count,
                          admin_users=admin_users,
                          recent_activities=recent_activities)

# User management routes
@bp.route('/users')
@login_required
@admin_required
def list_users():
    """List all users"""
    page = request.args.get('page', 1, type=int)
    per_page = current_app.config['USERS_PER_PAGE']
    status_filter = request.args.get('status', '').strip()

    # Build query based on status filter
    query = User.query
    if status_filter == 'active':
        query = query.filter_by(is_active=True)
    elif status_filter == 'inactive':
        query = query.filter_by(is_active=False)

    users = query.order_by(User.username).paginate(
        page=page, per_page=per_page, error_out=False)

    return render_template('admin/users/list.html', users=users, status_filter=status_filter)

@bp.route('/users/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_user():
    """Create a new user"""
    if request.method == 'POST':
        # Get form data
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        department = request.form.get('department', '').strip()
        phone = request.form.get('phone', '').strip()
        bio = request.form.get('bio', '').strip()
        role_id = request.form.get('role_id', type=int)
        is_active = request.form.get('is_active') == 'on'
        is_admin = request.form.get('is_admin') == 'on'

        # Validation
        errors = []

        if not username:
            errors.append('اسم المستخدم مطلوب')
        elif User.query.filter_by(username=username).first():
            errors.append('اسم المستخدم موجود بالفعل')

        if not email:
            errors.append('البريد الإلكتروني مطلوب')
        elif User.query.filter_by(email=email).first():
            errors.append('البريد الإلكتروني موجود بالفعل')

        if not password:
            errors.append('كلمة المرور مطلوبة')
        elif len(password) < 6:
            errors.append('كلمة المرور يجب أن تكون على الأقل 6 أحرف')
        elif password != confirm_password:
            errors.append('كلمات المرور غير متطابقة')

        if not role_id:
            errors.append('يجب تحديد دور للمستخدم')

        if errors:
            for error in errors:
                flash(error, 'error')
            roles = Role.query.all()
            return render_template('admin/users/create.html', roles=roles)

        try:
            # Create new user
            user = User(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                department=department,
                phone=phone,
                bio=bio,
                is_active=is_active,
                is_admin=is_admin
            )

            # Set password
            user.set_password(password)

            # Set role
            role = Role.query.get(role_id)
            if role:
                user.role = role

            # Set timestamps
            user.created_at = datetime.now()
            user.updated_at = datetime.now()

            # Add to database
            db.session.add(user)
            db.session.commit()

            # Log the action
            try:
                from app.models.audit import AuditLog, AuditAction
                AuditLog.log_user_action(
                    action=AuditAction.USER_CREATE,
                    target_user=user,
                    user=current_user,
                    description=f'Created new user: {user.username}',
                    ip_address=request.remote_addr,
                    user_agent=request.headers.get('User-Agent')
                )
            except Exception as e:
                # Don't fail user creation if audit logging fails
                current_app.logger.warning(f'Failed to log user creation: {str(e)}')

            flash(f'تم إنشاء المستخدم {username} بنجاح', 'success')
            return redirect(url_for('admin.list_users'))

        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في إنشاء المستخدم: {str(e)}', 'error')
            roles = Role.query.all()
            return render_template('admin/users/create.html', roles=roles)

    # Show create form
    roles = Role.query.all()
    return render_template('admin/users/create.html', roles=roles)

@bp.route('/users/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(id):
    """Edit an existing user"""
    user = User.query.get_or_404(id)

    if request.method == 'POST':
        # Get form data
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        department = request.form.get('department', '').strip()
        phone = request.form.get('phone', '').strip()
        bio = request.form.get('bio', '').strip()
        role_id = request.form.get('role_id', type=int)
        is_active = request.form.get('is_active') == 'on'
        is_admin = request.form.get('is_admin') == 'on'

        # Validation
        errors = []

        if not email:
            errors.append('البريد الإلكتروني مطلوب')
        else:
            # Check if email is already used by another user
            existing_user = User.query.filter_by(email=email).first()
            if existing_user and existing_user.id != user.id:
                errors.append('البريد الإلكتروني موجود بالفعل')

        # Validate password if provided
        if password:
            if len(password) < 6:
                errors.append('كلمة المرور يجب أن تكون على الأقل 6 أحرف')
            elif password != confirm_password:
                errors.append('كلمات المرور غير متطابقة')

        if not role_id:
            errors.append('يجب تحديد دور للمستخدم')

        if errors:
            for error in errors:
                flash(error, 'error')
            roles = Role.query.all()
            return render_template('admin/users/edit.html', user=user, roles=roles)

        try:
            # Update user data
            user.email = email
            user.first_name = first_name
            user.last_name = last_name
            user.department = department
            user.phone = phone
            user.bio = bio
            user.updated_at = datetime.now()

            # Update password if provided
            if password:
                user.set_password(password)

            # Update role (but protect super admin)
            if user.role and user.role.name != 'super_admin':
                role = Role.query.get(role_id)
                if role:
                    user.role = role

            # Update status (but protect super admin)
            if user.role and user.role.name != 'super_admin':
                user.is_active = is_active
                user.is_admin = is_admin

            # Commit changes
            db.session.commit()

            # Log the action
            try:
                from app.models.audit import AuditLog, AuditAction
                AuditLog.log_user_action(
                    action=AuditAction.USER_UPDATE,
                    target_user=user,
                    user=current_user,
                    description=f'Updated user: {user.username}',
                    ip_address=request.remote_addr,
                    user_agent=request.headers.get('User-Agent')
                )
            except Exception as e:
                # Don't fail user update if audit logging fails
                current_app.logger.warning(f'Failed to log user update: {str(e)}')

            flash(f'تم تحديث بيانات المستخدم {user.username} بنجاح', 'success')
            return redirect(url_for('admin.list_users'))

        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في تحديث المستخدم: {str(e)}', 'error')
            roles = Role.query.all()
            return render_template('admin/users/edit.html', user=user, roles=roles)

    # Show edit form
    roles = Role.query.all()
    return render_template('admin/users/edit.html', user=user, roles=roles)

@bp.route('/users/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_user(id):
    """Delete a user"""
    user = User.query.get_or_404(id)
    
    # Prevent deleting yourself
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('admin.list_users'))
    
    db.session.delete(user)
    db.session.commit()
    
    flash('User has been deleted.', 'success')
    return redirect(url_for('admin.list_users'))

# Category management routes
@bp.route('/categories')
@login_required
@admin_required
def list_categories():
    """List all categories in hierarchical structure"""
    from sqlalchemy import func



    def build_category_tree(parent_id=None, level=0):
        """Build hierarchical category tree structure with custom sorting"""
        categories = Category.query.filter_by(parent_id=parent_id).all()

        # Custom sorting: categories with children first (by creation date), then categories without children
        def sort_categories(cat):
            has_children = Category.query.filter_by(parent_id=cat.id).count() > 0
            # Return tuple: (has_children_inverted, created_at)
            # has_children_inverted = False for categories with children (they come first)
            return (not has_children, cat.created_at or datetime.min)

        categories.sort(key=sort_categories)

        result = []
        for category in categories:
            # Get subcategories recursively
            children = build_category_tree(category.id, level + 1)

            # Convert to dict for JSON serialization
            category_dict = {
                'id': category.id,
                'name': category.name,
                'name_ar': category.name_ar,
                'description': category.description,
                'description_ar': category.description_ar,
                'icon': category.icon,
                'color': category.color,
                'color_hex': category.get_color_hex(),
                'color_class': category.get_color_class(),
                'level': level,
                'parent_id': category.parent_id,
                'total_documents': get_category_document_count(category.id),
                'direct_documents': category.documents.count() if category.documents else 0,
                'hierarchy_path': get_category_path(category),
                'created_at': category.created_at.isoformat() if category.created_at else None,
                'created_at_formatted': category.created_at.strftime('%Y-%m-%d %H:%M') if category.created_at else 'غير محدد',
                'children': children,
                'has_children': len(children) > 0
            }

            result.append(category_dict)

        return result

    def get_category_path(category):
        """Get full path of category hierarchy"""
        path = []
        current = category
        while current:
            path.insert(0, current.get_display_name())
            current = current.parent
        return ' ← '.join(path)

    # Get return_to parameter to determine navigation context
    return_to = request.args.get('return_to', 'admin')

    # Build hierarchical categories tree
    categories_tree = build_category_tree()



    # Calculate detailed statistics
    def calculate_statistics():
        """Calculate comprehensive category statistics"""
        all_categories = Category.query.all()

        stats = {
            'total_categories': len(all_categories),
            'main_categories': 0,
            'subcategories_by_level': {},
            'total_documents': 0,
            'documents_by_level': {},
            'categories_with_documents': 0,
            'empty_categories': 0,
            'deepest_level': 0,
            'main_categories_chart_data': []
        }

        # Initialize level counters
        for i in range(6):  # Support up to level 5
            stats['subcategories_by_level'][i] = 0
            stats['documents_by_level'][i] = 0

        for category in all_categories:
            level = get_category_level(category)

            # Update deepest level
            if level > stats['deepest_level']:
                stats['deepest_level'] = level

            # Count categories by level
            if level == 0:
                stats['main_categories'] += 1
            else:
                stats['subcategories_by_level'][level] += 1

            # Count documents
            direct_docs = category.documents.count() if category.documents else 0
            total_docs = get_category_document_count(category.id)

            stats['total_documents'] += direct_docs
            stats['documents_by_level'][level] += direct_docs

            # Count categories with/without documents (only subcategories, not main categories)
            if level > 0:  # Only count subcategories (level 1 and above)
                if total_docs > 0:
                    stats['categories_with_documents'] += 1
                else:
                    stats['empty_categories'] += 1

        # Prepare chart data for main categories
        main_categories = Category.query.filter_by(parent_id=None).all()
        for main_cat in main_categories:
            total_docs = get_category_document_count(main_cat.id)
            stats['main_categories_chart_data'].append({
                'name': main_cat.name_ar or main_cat.name,
                'documents': total_docs,
                'color': main_cat.color or '#007bff'
            })

        # Sort by document count (descending)
        stats['main_categories_chart_data'].sort(key=lambda x: x['documents'], reverse=True)

        return stats

    statistics = calculate_statistics()

    return render_template('admin/categories/list.html',
                         categories=categories_tree,
                         statistics=statistics,
                         return_to=return_to)

@bp.route('/categories/subcategory-details/<int:id>')
@login_required
@admin_required
def subcategory_details(id):
    """Get subcategory details for a main category - includes all levels"""
    try:
        category = Category.query.get_or_404(id)

        def get_all_subcategories_recursive(parent_id, level=1):
            """Get all subcategories recursively"""
            subcategories = Category.query.filter_by(parent_id=parent_id).all()
            all_subcats = []

            for subcategory in subcategories:
                # Add current subcategory
                all_subcats.append({
                    'category': subcategory,
                    'level': level
                })

                # Add all children recursively
                children = get_all_subcategories_recursive(subcategory.id, level + 1)
                all_subcats.extend(children)

            return all_subcats

        # Get all subcategories at all levels
        all_subcategories = get_all_subcategories_recursive(id)

        subcategory_data = []
        total_documents = 0

        for subcat_info in all_subcategories:
            subcategory = subcat_info['category']
            level = subcat_info['level']

            # Count documents in this subcategory and all its children
            doc_count = get_category_document_count(subcategory.id)
            total_documents += doc_count

            # Build category path
            path_parts = []
            current = subcategory.parent
            while current and current.id != id:  # Stop at the main category
                path_parts.insert(0, current.name_ar or current.name)
                current = current.parent

            # Create level indicator
            level_indicator = '└─ ' * (level - 1) if level > 1 else ''
            display_name = level_indicator + (subcategory.name_ar or subcategory.name)

            subcategory_data.append({
                'id': subcategory.id,
                'name': display_name,
                'original_name': subcategory.name_ar or subcategory.name,
                'path': ' ← '.join(path_parts) if path_parts else None,
                'documents': doc_count,
                'color': subcategory.color,
                'color_hex': subcategory.get_color_hex(),
                'color_class': subcategory.get_color_class(),
                'level': level,
                'is_direct_child': level == 1
            })

        # Sort by level first, then by document count (descending)
        subcategory_data.sort(key=lambda x: (x['level'], -x['documents']))

        return jsonify({
            'success': True,
            'subcategories': subcategory_data,
            'total_documents': total_documents,
            'category_name': category.name_ar or category.name
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@bp.route('/categories/details/<int:id>')
@login_required
@admin_required
def category_details(id):
    """Get detailed information about a category"""
    from sqlalchemy import func

    try:
        category = Category.query.get_or_404(id)

        def get_category_document_count(category_id):
            """Get total document count for a category including all subcategories recursively"""
            direct_count = db.session.query(func.count(Document.id))\
                .filter(Document.category_id == category_id)\
                .scalar() or 0

            subcategories = Category.query.filter_by(parent_id=category_id).all()
            recursive_count = 0
            for subcat in subcategories:
                recursive_count += get_category_document_count(subcat.id)

            return direct_count + recursive_count

        def get_category_path(category):
            """Get full path of category hierarchy"""
            path = []
            current = category
            while current:
                path.insert(0, current.get_display_name())
                current = current.parent
            return ' ← '.join(path)

        def get_category_level(category):
            """Get category level in hierarchy"""
            level = 0
            current = category.parent
            while current:
                level += 1
                current = current.parent
            return level

        def build_full_category_tree(target_category):
            """Build complete category tree showing the target category and all its relationships"""
            # Get root category
            root = target_category
            while root.parent:
                root = root.parent

            def build_tree_recursive(cat, target_id, current_level=0):
                children = Category.query.filter_by(parent_id=cat.id).all()

                # Custom sorting: categories with children first, then by creation date
                def sort_categories(category):
                    has_children = Category.query.filter_by(parent_id=category.id).count() > 0
                    return (not has_children, category.created_at or datetime.min)

                children.sort(key=sort_categories)

                children_data = []
                for child in children:
                    child_data = {
                        'id': child.id,
                        'name': child.name,
                        'name_ar': child.name_ar,
                        'icon': child.icon,
                        'color': child.color,
                        'color_hex': child.get_color_hex(),
                        'color_class': child.get_color_class(),
                        'level': current_level + 1,
                        'total_documents': get_category_document_count(child.id),
                        'direct_documents': child.documents.count() if child.documents else 0,
                        'created_at': child.created_at.strftime('%Y-%m-%d %H:%M') if child.created_at else 'غير محدد',
                        'is_target': child.id == target_id,
                        'children': build_tree_recursive(child, target_id, current_level + 1)
                    }
                    children_data.append(child_data)

                return children_data

            # Build the complete tree starting from root
            root_data = {
                'id': root.id,
                'name': root.name,
                'name_ar': root.name_ar,
                'icon': root.icon,
                'color': root.color,
                'color_hex': root.get_color_hex(),
                'color_class': root.get_color_class(),
                'level': 0,
                'total_documents': get_category_document_count(root.id),
                'direct_documents': root.documents.count() if root.documents else 0,
                'created_at': root.created_at.strftime('%Y-%m-%d %H:%M') if root.created_at else 'غير محدد',
                'is_target': root.id == target_category.id,
                'children': build_tree_recursive(root, target_category.id, 0)
            }

            return root_data

        # Build category details with complete tree
        category_tree = build_full_category_tree(category)

        category_data = {
            'id': category.id,
            'name': category.name,
            'name_ar': category.name_ar,
            'description': category.description,
            'description_ar': category.description_ar,
            'icon': category.icon,
            'color': category.color,
            'color_hex': category.get_color_hex(),
            'color_class': category.get_color_class(),
            'hierarchy_path': get_category_path(category),
            'level': get_category_level(category),
            'total_documents': get_category_document_count(category.id),
            'direct_documents': category.documents.count() if category.documents else 0,
            'subcategories_count': Category.query.filter_by(parent_id=category.id).count(),
            'created_at': category.created_at.strftime('%Y-%m-%d %H:%M') if category.created_at else None,
            'updated_at': category.updated_at.strftime('%Y-%m-%d %H:%M') if category.updated_at else None,
            'full_tree': category_tree
        }

        return jsonify({
            'success': True,
            'category': category_data
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@bp.route('/categories/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_category():
    """Create a new category"""
    # Get main categories (categories without parent)
    main_categories = Category.query.filter_by(parent_id=None).all()

    # Check if parent is specified in URL
    suggested_parent_id = request.args.get('parent_id', type=int)

    # Get return_to parameter for redirect after save
    return_to = request.args.get('return_to', 'admin')

    # Get parent category info if specified
    parent_category = None
    parent_hierarchy = []
    if suggested_parent_id:
        parent_category = Category.query.get(suggested_parent_id)
        if parent_category:
            # Build hierarchy path for the parent
            current = parent_category
            while current:
                parent_hierarchy.insert(0, {
                    'id': current.id,
                    'name': current.name_ar or current.name,
                    'level': get_category_level(current)
                })
                current = current.parent

    if request.method == 'POST':
        # Get form data
        name_ar = request.form.get('name_ar', '').strip()
        name = request.form.get('name', '').strip()
        description_ar = request.form.get('description_ar', '').strip()
        description = request.form.get('description', '').strip()
        parent_id = request.form.get('parent_id', type=int)
        main_category_id = request.form.get('main_category_id', '').strip()
        icon = request.form.get('icon', 'folder').strip()
        color = request.form.get('color', '#007bff').strip()

        # Get return_to from form data (takes precedence over URL parameter)
        return_to = request.form.get('return_to', return_to)

        # Validation
        errors = []

        if not name_ar:
            errors.append('اسم الفئة بالعربية مطلوب')

        # التحقق من اختيار الفئة الرئيسية
        if not main_category_id:
            errors.append('يجب اختيار الفئة الرئيسية أو إنشاء فئة رئيسية جديدة')

        # إذا لم يكن الخيار "إنشاء فئة رئيسية جديدة" ولم يتم تحديد فئة رئيسية
        if main_category_id != 'new_main' and not parent_id:
            errors.append('يجب تحديد فئة رئيسية للفئة الفرعية')

        # Check if parent exists (if provided)
        if parent_id:
            parent_category = Category.query.get(parent_id)
            if not parent_category:
                errors.append('الفئة الرئيسية المحددة غير موجودة')

        if errors:
            for error in errors:
                flash(error, 'danger')
            return render_template('admin/categories/create.html',
                                 main_categories=main_categories,
                                 form_data=request.form,
                                 return_to=return_to)

        # Create new category
        # إذا كان الخيار "إنشاء فئة رئيسية جديدة" فلا تحدد فئة رئيسية
        final_parent_id = None
        if main_category_id != 'new_main' and parent_id:
            final_parent_id = parent_id

        category = Category(
            name=name if name else name_ar,  # Use Arabic name as fallback if English name is empty
            name_ar=name_ar,
            description=description if description else None,
            description_ar=description_ar if description_ar else None,
            icon=icon,
            color=color,
            parent_id=final_parent_id
        )

        db.session.add(category)
        db.session.commit()

        flash('تم إنشاء الفئة بنجاح', 'success')

        # Always redirect to admin categories
        return redirect(url_for('admin.list_categories'))

    return render_template('admin/categories/create.html',
                         main_categories=main_categories,
                         suggested_parent_id=suggested_parent_id,
                         parent_category=parent_category,
                         parent_hierarchy=parent_hierarchy,
                         return_to=return_to)

@bp.route('/categories/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_category(id):
    """Edit an existing category"""
    category = Category.query.get_or_404(id)

    # Get return_to parameter for redirect after save
    return_to = request.args.get('return_to', 'admin')

    if request.method == 'POST':
        # Get form data
        name_ar = request.form.get('name_ar', '').strip()
        name = request.form.get('name', '').strip()
        description_ar = request.form.get('description_ar', '').strip()
        description = request.form.get('description', '').strip()
        parent_id = request.form.get('parent_id', type=int)
        icon = request.form.get('icon', 'folder').strip()
        color = request.form.get('color', '#007bff').strip()

        # Get return_to from form data (takes precedence over URL parameter)
        return_to = request.form.get('return_to', return_to)

        # Validation
        errors = []

        if not name_ar:
            errors.append('اسم الفئة بالعربية مطلوب')

        # Check if parent is not the category itself or its child
        if parent_id and parent_id == category.id:
            errors.append('لا يمكن أن تكون الفئة رئيسية لنفسها')

        if errors:
            for error in errors:
                flash(error, 'error')
            return render_template('admin/categories/edit.html', category=category, return_to=return_to)

        try:
            # Update category data
            category.name_ar = name_ar
            # If English name is empty, use Arabic name as fallback
            category.name = name if name else name_ar
            category.description_ar = description_ar if description_ar else None
            category.description = description if description else None
            category.icon = icon
            category.color = color
            category.updated_at = datetime.now()

            # Update parent
            if parent_id and parent_id > 0:
                parent_category = Category.query.get(parent_id)
                if parent_category:
                    category.parent_id = parent_id
                else:
                    category.parent_id = None
            else:
                category.parent_id = None

            # Commit changes
            db.session.commit()

            flash(f'تم تحديث الفئة {name_ar} بنجاح', 'success')

            # Always redirect to admin categories
            return redirect(url_for('admin.list_categories'))

        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في تحديث الفئة: {str(e)}', 'error')
            return render_template('admin/categories/edit.html', category=category, return_to=return_to)

    # Get all main categories for the hierarchical form
    main_categories = Category.query.filter_by(parent_id=None).all()

    # Build current category hierarchy
    current_hierarchy = []
    if category.parent_id:
        current = category.parent
        while current:
            current_hierarchy.insert(0, {
                'id': current.id,
                'name': current.name_ar or current.name,
                'level': get_category_level(current)
            })
            current = current.parent

    # Check if category is main category (no parent)
    is_main_category = category.parent_id is None

    # Check if category has documents
    has_documents = category.documents.count() > 0

    # Check if category can be made main category
    # Main categories should not contain documents directly
    can_be_main_category = not has_documents

    # Show edit form
    return render_template('admin/categories/edit.html',
                         category=category,
                         main_categories=main_categories,
                         current_hierarchy=current_hierarchy,
                         is_main_category=is_main_category,
                         has_documents=has_documents,
                         can_be_main_category=can_be_main_category,
                         return_to=return_to)

@bp.route('/categories/delete/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def delete_category(id):
    """Delete a category and all its subcategories and documents"""
    from flask_wtf.csrf import validate_csrf
    from werkzeug.exceptions import BadRequest

    # Get return_to parameter to determine where to redirect after deletion
    return_to = request.args.get('return_to', 'admin')

    # Handle GET requests (direct access to delete URL)
    if request.method == 'GET':
        flash('لا يمكن حذف الفئة من خلال الرابط المباشر. يرجى استخدام زر الحذف في قائمة الفئات.', 'warning')
        return redirect(url_for('admin.list_categories'))

    # Check if CSRF token exists
    csrf_token = request.form.get('csrf_token')
    if not csrf_token:
        flash('رمز الأمان مفقود. يرجى استخدام النموذج المخصص للحذف.', 'error')
        return redirect(url_for('admin.list_categories'))

    try:
        # Validate CSRF token
        validate_csrf(csrf_token)
    except Exception as e:
        flash('رمز الأمان غير صحيح. يرجى المحاولة مرة أخرى.', 'error')
        return redirect(url_for('admin.list_categories'))

    category = Category.query.get_or_404(id)
    category_name = category.name_ar or category.name

    try:
        # Count subcategories and documents that will be deleted
        def count_all_subcategories_and_documents(cat):
            """Recursively count all subcategories and documents"""
            subcategory_count = 0
            document_count = cat.documents.count()

            for child in cat.children:
                child_sub_count, child_doc_count = count_all_subcategories_and_documents(child)
                subcategory_count += 1 + child_sub_count  # +1 for the child itself
                document_count += child_doc_count

            return subcategory_count, document_count

        subcategory_count, document_count = count_all_subcategories_and_documents(category)

        # Delete all documents in this category and its subcategories first
        def delete_documents_recursively(cat):
            """Recursively delete all documents in category and subcategories"""
            # Delete documents in current category
            for doc in cat.documents:
                # Delete physical file if exists
                if doc.file_path and os.path.exists(doc.file_path):
                    try:
                        os.remove(doc.file_path)
                    except:
                        pass  # Continue even if file deletion fails

                # Delete document versions
                for version in doc.versions:
                    if version.file_path and os.path.exists(version.file_path):
                        try:
                            os.remove(version.file_path)
                        except:
                            pass

                db.session.delete(doc)

            # Recursively delete documents in subcategories
            for child in cat.children:
                delete_documents_recursively(child)

        # Delete documents first
        delete_documents_recursively(category)

        # Now delete the category (this will cascade to subcategories due to the relationship)
        db.session.delete(category)
        db.session.commit()

        # Create success message with details
        message_parts = [f'تم حذف الفئة "{category_name}" بنجاح']
        if subcategory_count > 0:
            message_parts.append(f'تم حذف {subcategory_count} فئة فرعية')
        if document_count > 0:
            message_parts.append(f'تم حذف {document_count} وثيقة')

        flash('. '.join(message_parts) + '.', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في حذف الفئة: {str(e)}', 'error')

    # Always redirect to admin categories
    return redirect(url_for('admin.list_categories'))

# Tag management routes will be added later if needed

@bp.route('/categories/hierarchy/<int:id>')
@login_required
@admin_required
def category_hierarchy(id):
    """Get category hierarchy path via AJAX"""
    try:
        category = Category.query.get_or_404(id)

        # Build hierarchy from root to current category
        hierarchy = []
        current = category

        # Build hierarchy from current category up to root
        while current:
            hierarchy.insert(0, {
                'id': current.id,
                'name': current.name,
                'name_ar': current.name_ar,
                'level': get_category_level(current)
            })
            current = current.parent

        return jsonify({
            'success': True,
            'hierarchy': hierarchy
        })

    except Exception as e:
        current_app.logger.error(f"Error getting category hierarchy: {str(e)}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

# Audit log routes
@bp.route('/logs')
@login_required
@admin_required
def list_logs():
    """List audit logs with filtering"""
    page = request.args.get('page', 1, type=int)
    per_page = 50

    # Get filter parameters
    action_filter = request.args.get('action', '').strip()
    user_filter = request.args.get('user', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    success_filter = request.args.get('success', '').strip()

    # Build query
    query = AuditLog.query

    # Apply filters
    if action_filter:
        query = query.filter(AuditLog.action == action_filter)

    if user_filter:
        # Filter by user ID if it's a number, otherwise by username
        if user_filter.isdigit():
            query = query.filter(AuditLog.user_id == int(user_filter))
        else:
            query = query.filter(AuditLog.username.ilike(f'%{user_filter}%'))

    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(AuditLog.created_at >= from_date)
        except ValueError:
            flash('تاريخ البداية غير صحيح', 'error')

    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            # Add 23:59:59 to include the entire day
            to_date = to_date.replace(hour=23, minute=59, second=59)
            query = query.filter(AuditLog.created_at <= to_date)
        except ValueError:
            flash('تاريخ النهاية غير صحيح', 'error')

    if success_filter:
        if success_filter == 'true':
            query = query.filter(AuditLog.success == True)
        elif success_filter == 'false':
            query = query.filter(AuditLog.success == False)

    # Order and paginate
    logs = query.order_by(AuditLog.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False)

    # Get all users for dropdown
    users = User.query.filter_by(is_active=True).order_by(User.username).all()

    # Get available actions for dropdown
    available_actions_raw = db.session.query(AuditLog.action).distinct().order_by(AuditLog.action).all()
    available_actions = []
    for action in available_actions_raw:
        if action[0]:  # Make sure action is not None
            # Handle both string and enum types
            action_value = action[0] if isinstance(action[0], str) else action[0].name if hasattr(action[0], 'name') else str(action[0])
            available_actions.append(action_value)

    # Debug: Print available actions
    print("Available actions:", available_actions)

    # Create a list with both English values and Arabic translations
    from app.models.audit import AuditAction
    actions_with_translations = []
    for action in available_actions:
        arabic_text = AuditAction.get_action_description(action)
        actions_with_translations.append({
            'value': action,
            'text': arabic_text
        })

    print("Actions with translations:", actions_with_translations)

    # Create action translations dictionary
    action_translations = {
        'LOGIN': 'تسجيل دخول',
        'LOGOUT': 'تسجيل خروج',
        'LOGIN_FAILED': 'فشل تسجيل الدخول',
        'DOCUMENT_CREATE': 'إنشاء وثيقة',
        'DOCUMENT_UPDATE': 'تحديث وثيقة',
        'DOCUMENT_DELETE': 'حذف وثيقة',
        'DOCUMENT_VIEW': 'عرض وثيقة',
        'DOCUMENT_DOWNLOAD': 'تحميل وثيقة',
        'USER_CREATE': 'إنشاء مستخدم',
        'USER_UPDATE': 'تحديث مستخدم',
        'USER_DELETE': 'حذف مستخدم',
        'USER_ACTIVATE': 'تفعيل مستخدم',
        'USER_DEACTIVATE': 'إلغاء تفعيل مستخدم',
        'USER_ROLE_CHANGE': 'تغيير دور مستخدم',
        'CATEGORY_CREATE': 'إنشاء فئة',
        'CATEGORY_UPDATE': 'تحديث فئة',
        'CATEGORY_DELETE': 'حذف فئة',
        'ROLE_CREATE': 'إنشاء دور',
        'ROLE_UPDATE': 'تحديث دور',
        'ROLE_DELETE': 'حذف دور',
        'PERMISSION_UPDATE': 'تحديث صلاحية',
        'SYSTEM_BACKUP': 'نسخ احتياطي',
        'SYSTEM_RESTORE': 'استعادة نسخة احتياطية',
        'SETTINGS_UPDATE': 'تحديث الإعدادات'
    }

    # Prepare filter values for template
    filters = {
        'action': action_filter,
        'user': user_filter,
        'date_from': date_from,
        'date_to': date_to,
        'success': success_filter
    }

    return render_template('admin/logs/list.html',
                         logs=logs,
                         users=users,
                         available_actions=available_actions,
                         actions_with_translations=actions_with_translations,
                         filters=filters,
                         action_translations=action_translations,
                         translate_action=translate_action)

# User management with permissions
@bp.route('/users/<int:user_id>/permissions')
@login_required
@super_admin_required
def user_permissions(user_id):
    """View and manage user permissions"""
    user = User.query.get_or_404(user_id)
    all_roles = Role.query.all()
    all_permissions = Permission.query.all()

    return render_template('admin/user_permissions.html',
                          user=user,
                          all_roles=all_roles,
                          all_permissions=all_permissions,
                          translate_permission_name=translate_permission_name)

@bp.route('/users/<int:user_id>/update-role', methods=['POST'])
@login_required
@super_admin_required
def update_user_role(user_id):
    """Update user role"""
    user = User.query.get_or_404(user_id)

    # Prevent changing super admin role
    if user.role and user.role.name == 'super_admin' and current_user.id != user.id:
        flash('لا يمكن تغيير دور مدير النظام الرئيسي', 'danger')
        return redirect(url_for('admin.user_permissions', user_id=user.id))

    role_id = request.form.get('role_id')
    if role_id:
        role = Role.query.get(role_id)
        if role:
            user.role = role
            user.updated_at = datetime.now()

            try:
                db.session.commit()
                flash(f'تم تحديث دور المستخدم إلى: {role.description}', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'خطأ في تحديث الدور: {str(e)}', 'danger')

    return redirect(url_for('admin.user_permissions', user_id=user.id))

@bp.route('/users/<int:user_id>/toggle-status', methods=['POST'])
@login_required
@super_admin_required
def toggle_user_status(user_id):
    """Toggle user active status"""
    user = User.query.get_or_404(user_id)

    # Prevent deactivating super admin
    if user.role and user.role.name == 'super_admin':
        return json.dumps({'success': False, 'message': 'لا يمكن إلغاء تفعيل مدير النظام الرئيسي'})

    user.is_active = not user.is_active
    user.updated_at = datetime.now()

    try:
        db.session.commit()
        status = 'مفعل' if user.is_active else 'معطل'
        return json.dumps({'success': True, 'message': f'تم {status} المستخدم بنجاح', 'is_active': user.is_active})
    except Exception as e:
        db.session.rollback()
        return json.dumps({'success': False, 'message': f'خطأ: {str(e)}'})

@bp.route('/users/<int:user_id>/toggle-admin', methods=['POST'])
@login_required
@super_admin_required
def toggle_user_admin(user_id):
    """Toggle user admin status"""
    user = User.query.get_or_404(user_id)

    # Prevent removing admin from super admin
    if user.role and user.role.name == 'super_admin':
        return json.dumps({'success': False, 'message': 'لا يمكن إزالة صلاحية الإدارة من مدير النظام الرئيسي'})

    user.is_admin = not user.is_admin
    user.updated_at = datetime.now()

    # When granting admin status, grant all individual permissions
    if user.is_admin:
        user.can_access_admin = True
        user.can_manage_users = True
        user.can_manage_categories = True
        user.can_delete_documents = True
        user.can_view_confidential = True
        user.can_view_audit_logs = True
        user.can_manage_system = True
        user.can_view_stats = True
    else:
        # When removing admin status, remove all individual permissions
        user.can_access_admin = False
        user.can_manage_users = False
        user.can_manage_categories = False
        user.can_delete_documents = False
        user.can_view_confidential = False
        user.can_view_audit_logs = False
        user.can_manage_system = False
        user.can_view_stats = False

    try:
        db.session.commit()
        status = 'مدير' if user.is_admin else 'مستخدم عادي'
        return json.dumps({'success': True, 'message': f'تم تحديث المستخدم إلى: {status}', 'is_admin': user.is_admin})
    except Exception as e:
        db.session.rollback()
        return json.dumps({'success': False, 'message': f'خطأ: {str(e)}'})

@bp.route('/users/<int:user_id>/update-admin-permissions', methods=['POST'])
@login_required
@super_admin_required
def update_admin_permissions(user_id):
    """Update individual admin permissions"""
    from flask import Response

    try:
        print(f"Received request for user_id: {user_id}")  # للتشخيص
        print(f"Request content type: {request.content_type}")  # للتشخيص
        print(f"Request data: {request.data}")  # للتشخيص

        user = User.query.get_or_404(user_id)

        # Prevent editing super admin permissions
        if user.role and user.role.name == 'super_admin':
            response_data = json.dumps({'success': False, 'message': 'لا يمكن تعديل صلاحيات مدير النظام الرئيسي'}, ensure_ascii=False)
            return Response(response_data, content_type='application/json; charset=utf-8')

        # Get permission data from request
        permissions = request.get_json()
        print(f"Parsed permissions: {permissions}")  # للتشخيص

        if not permissions:
            response_data = json.dumps({'success': False, 'message': 'لم يتم استلام بيانات الصلاحيات'}, ensure_ascii=False)
            return Response(response_data, content_type='application/json; charset=utf-8')

        # Update individual permissions
        user.can_access_admin = permissions.get('can_access_admin', False)
        user.can_manage_users = permissions.get('can_manage_users', False)
        user.can_manage_categories = permissions.get('can_manage_categories', False)
        user.can_delete_documents = permissions.get('can_delete_documents', False)
        user.can_view_confidential = permissions.get('can_view_confidential', False)
        user.can_view_audit_logs = permissions.get('can_view_audit_logs', False)
        user.can_manage_system = permissions.get('can_manage_system', False)
        user.can_view_stats = permissions.get('can_view_stats', False)

        # Update is_admin based on whether any admin permission is granted
        user.is_admin = any([
            user.can_access_admin, user.can_manage_users, user.can_manage_categories,
            user.can_delete_documents, user.can_view_confidential, user.can_view_audit_logs,
            user.can_manage_system, user.can_view_stats
        ])

        user.updated_at = datetime.now()
        db.session.commit()

        print(f"Successfully updated permissions for user: {user.username}")  # للتشخيص

        response_data = json.dumps({
            'success': True,
            'message': 'تم تحديث الصلاحيات بنجاح',
            'is_admin': user.is_admin
        }, ensure_ascii=False)
        return Response(response_data, content_type='application/json; charset=utf-8')

    except Exception as e:
        db.session.rollback()
        print(f"Error updating admin permissions: {str(e)}")  # للتشخيص
        import traceback
        traceback.print_exc()  # للتشخيص المفصل

        response_data = json.dumps({'success': False, 'message': f'خطأ في الخادم: {str(e)}'}, ensure_ascii=False)
        return Response(response_data, content_type='application/json; charset=utf-8', status=500)

# Role management
@bp.route('/roles')
@login_required
@super_admin_required
def roles():
    """List all roles"""
    roles = Role.query.all()
    return render_template('admin/roles.html',
                         roles=roles,
                         translate_role_name=translate_role_name,
                         translate_permission_name=translate_permission_name)

@bp.route('/roles/<int:role_id>')
@login_required
@super_admin_required
def role_detail(role_id):
    """View role details"""
    role = Role.query.get_or_404(role_id)
    users_with_role = User.query.filter_by(role_id=role.id).all()
    return render_template('admin/role_detail.html', role=role, users_with_role=users_with_role)

@bp.route('/roles/<int:role_id>/edit', methods=['GET', 'POST'])
@login_required
@super_admin_required
def edit_role(role_id):
    """Edit role permissions"""
    role = Role.query.get_or_404(role_id)
    permissions = Permission.query.all()

    if request.method == 'POST':
        # Prevent editing super admin role
        if role.name == 'super_admin':
            flash('لا يمكن تعديل دور مدير النظام الرئيسي', 'danger')
            return redirect(url_for('admin.role_detail', role_id=role.id))

        # Update role basic info
        role.description = request.form.get('description', role.description)

        # Update permissions
        selected_permissions = request.form.getlist('permissions')
        role.permissions = Permission.query.filter(Permission.id.in_(selected_permissions)).all()

        try:
            db.session.commit()
            flash('تم تحديث الدور بنجاح', 'success')
            return redirect(url_for('admin.role_detail', role_id=role.id))
        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في تحديث الدور: {str(e)}', 'danger')

    return render_template('admin/edit_role.html', role=role, permissions=permissions)

# Permission management
def translate_permission_name(permission_name):
    """Translate permission name to Arabic"""
    translations = {
        # Document permissions
        'view_documents': 'عرض الوثائق',
        'create_documents': 'إنشاء الوثائق',
        'edit_documents': 'تعديل الوثائق',
        'delete_documents': 'حذف الوثائق',
        'approve_documents': 'اعتماد الوثائق',
        'view_confidential_documents': 'عرض الوثائق السرية',
        'download_documents': 'تحميل الوثائق',
        'manage_document_versions': 'إدارة إصدارات الوثائق',

        # Category permissions
        'manage_categories': 'إدارة الفئات',
        'create_categories': 'إنشاء الفئات',
        'edit_categories': 'تعديل الفئات',
        'delete_categories': 'حذف الفئات',

        # User permissions
        'manage_users': 'إدارة المستخدمين',
        'create_users': 'إنشاء المستخدمين',
        'edit_users': 'تعديل المستخدمين',
        'delete_users': 'حذف المستخدمين',
        'view_users': 'عرض المستخدمين',
        'manage_user_roles': 'إدارة أدوار المستخدمين',

        # Role permissions
        'manage_roles': 'إدارة الأدوار',
        'create_roles': 'إنشاء الأدوار',
        'edit_roles': 'تعديل الأدوار',
        'delete_roles': 'حذف الأدوار',
        'assign_roles': 'تخصيص الأدوار',

        # System permissions
        'system_admin': 'إدارة النظام',
        'view_audit_logs': 'عرض سجل الأنشطة',
        'manage_system_settings': 'إدارة إعدادات النظام',
        'backup_system': 'نسخ احتياطي للنظام',
        'restore_system': 'استعادة النظام',
        'view_system_stats': 'عرض إحصائيات النظام',

        # Admin permissions
        'admin_access': 'الوصول للوحة الإدارة',
        'super_admin': 'مدير النظام الرئيسي',
        'manage_permissions': 'إدارة الصلاحيات'
    }
    return translations.get(permission_name, permission_name)

def translate_role_name(role_name):
    """Translate role name to Arabic"""
    translations = {
        'super_admin': 'مدير النظام الرئيسي',
        'admin': 'مدير النظام',
        'manager': 'مشرف',
        'supervisor': 'مراقب',
        'editor': 'محرر',
        'user': 'مستخدم عادي',
        'viewer': 'مشاهد',
        'guest': 'ضيف'
    }
    return translations.get(role_name, role_name)

@bp.route('/permissions')
@login_required
@super_admin_required
def permissions():
    """List all permissions"""
    permissions = Permission.query.all()
    roles = Role.query.all()
    return render_template('admin/permissions.html',
                         permissions=permissions,
                         roles=roles,
                         translate_permission_name=translate_permission_name)

# Settings management
@bp.route('/settings')
@login_required
@super_admin_required
def settings():
    """System settings"""
    return render_template('admin/settings.html')

# Performance Reports
@bp.route('/performance-reports')
@login_required
@admin_required
def performance_reports():
    """Performance monitoring and reports"""
    try:
        from datetime import datetime, timedelta
        from app.models.document import Document
        from app.models.audit import AuditLog

        # Get performance metrics for the last 30 days
        thirty_days_ago = datetime.now() - timedelta(days=30)
        seven_days_ago = datetime.now() - timedelta(days=7)
        today = datetime.now().date()

        # System performance metrics
        total_users = User.query.count()
        active_users = User.query.filter_by(is_active=True).count()
        total_documents = Document.query.count()
        total_categories = Category.query.count()

        # Activity metrics
        total_activities = AuditLog.query.count()
        recent_activities = AuditLog.query.filter(AuditLog.created_at >= seven_days_ago).count()
        today_activities = AuditLog.query.filter(AuditLog.created_at >= today).count()

        # Document metrics
        recent_documents = Document.query.filter(Document.created_at >= seven_days_ago).count()
        today_documents = Document.query.filter(Document.created_at >= today).count()

        # User activity metrics
        active_users_week = db.session.query(AuditLog.user_id).filter(
            AuditLog.created_at >= seven_days_ago
        ).distinct().count()

        # Calculate performance ratios
        user_activity_ratio = (active_users_week / total_users * 100) if total_users > 0 else 0
        document_growth_ratio = (recent_documents / total_documents * 100) if total_documents > 0 else 0

        # Get daily activity data for charts
        daily_activities = []
        for i in range(7):
            day = datetime.now().date() - timedelta(days=i)
            day_start = datetime.combine(day, datetime.min.time())
            day_end = datetime.combine(day, datetime.max.time())

            activities_count = AuditLog.query.filter(
                AuditLog.created_at >= day_start,
                AuditLog.created_at <= day_end
            ).count()

            daily_activities.append({
                'date': day.strftime('%Y-%m-%d'),
                'date_ar': day.strftime('%d/%m'),
                'activities': activities_count
            })

        daily_activities.reverse()  # Show oldest to newest

        # Get most active users
        most_active_users = db.session.query(
            AuditLog.username,
            db.func.count(AuditLog.id).label('activity_count')
        ).filter(
            AuditLog.created_at >= seven_days_ago
        ).group_by(AuditLog.username).order_by(
            db.func.count(AuditLog.id).desc()
        ).limit(5).all()

        # Get most common actions
        most_common_actions = db.session.query(
            AuditLog.action,
            db.func.count(AuditLog.id).label('action_count')
        ).filter(
            AuditLog.created_at >= seven_days_ago
        ).group_by(AuditLog.action).order_by(
            db.func.count(AuditLog.id).desc()
        ).limit(5).all()

        performance_data = {
            'system_metrics': {
                'total_users': total_users,
                'active_users': active_users,
                'total_documents': total_documents,
                'total_categories': total_categories,
                'user_activity_ratio': round(user_activity_ratio, 1),
                'document_growth_ratio': round(document_growth_ratio, 1)
            },
            'activity_metrics': {
                'total_activities': total_activities,
                'recent_activities': recent_activities,
                'today_activities': today_activities,
                'active_users_week': active_users_week
            },
            'daily_activities': daily_activities,
            'most_active_users': most_active_users,
            'most_common_actions': most_common_actions
        }

        return render_template('admin/performance_reports.html',
                             performance_data=performance_data,
                             translate_action=translate_action)

    except Exception as e:
        flash(f'خطأ في تحميل تقارير الأداء: {str(e)}', 'error')
        return redirect(url_for('admin.index'))

# Data Analytics
@bp.route('/data-analytics')
@login_required
@admin_required
def data_analytics():
    """Data analytics and usage analysis"""
    try:
        from datetime import datetime, timedelta
        from app.models.document import Document
        from app.models.audit import AuditLog

        # Time periods
        thirty_days_ago = datetime.now() - timedelta(days=30)
        seven_days_ago = datetime.now() - timedelta(days=7)

        # User analytics
        user_registration_trend = []
        for i in range(30):
            day = datetime.now().date() - timedelta(days=i)
            day_start = datetime.combine(day, datetime.min.time())
            day_end = datetime.combine(day, datetime.max.time())

            new_users = User.query.filter(
                User.created_at >= day_start,
                User.created_at <= day_end
            ).count()

            user_registration_trend.append({
                'date': day.strftime('%Y-%m-%d'),
                'date_ar': day.strftime('%d/%m'),
                'new_users': new_users
            })

        user_registration_trend.reverse()

        # Document analytics
        document_creation_trend = []
        for i in range(30):
            day = datetime.now().date() - timedelta(days=i)
            day_start = datetime.combine(day, datetime.min.time())
            day_end = datetime.combine(day, datetime.max.time())

            new_documents = Document.query.filter(
                Document.created_at >= day_start,
                Document.created_at <= day_end
            ).count()

            document_creation_trend.append({
                'date': day.strftime('%Y-%m-%d'),
                'date_ar': day.strftime('%d/%m'),
                'new_documents': new_documents
            })

        document_creation_trend.reverse()

        # Category usage analytics
        category_usage = db.session.query(
            Category.name,
            Category.name_ar,
            db.func.count(Document.id).label('document_count')
        ).outerjoin(Document).group_by(Category.id).order_by(
            db.func.count(Document.id).desc()
        ).limit(10).all()

        # User activity patterns
        hourly_activity = []
        for hour in range(24):
            activity_count = AuditLog.query.filter(
                AuditLog.created_at >= seven_days_ago,
                db.extract('hour', AuditLog.created_at) == hour
            ).count()

            hourly_activity.append({
                'hour': hour,
                'hour_display': f"{hour:02d}:00",
                'activity_count': activity_count
            })

        # Document type analytics
        document_types = db.session.query(
            Document.file_type,
            db.func.count(Document.id).label('count')
        ).group_by(Document.file_type).order_by(
            db.func.count(Document.id).desc()
        ).all()

        # User role distribution
        role_distribution = db.session.query(
            Role.name,
            Role.description,
            db.func.count(User.id).label('user_count')
        ).outerjoin(User).group_by(Role.id).all()

        analytics_data = {
            'user_registration_trend': user_registration_trend,
            'document_creation_trend': document_creation_trend,
            'category_usage': category_usage,
            'hourly_activity': hourly_activity,
            'document_types': document_types,
            'role_distribution': role_distribution
        }

        return render_template('admin/data_analytics.html',
                             analytics_data=analytics_data)

    except Exception as e:
        flash(f'خطأ في تحميل تحليل البيانات: {str(e)}', 'error')
        return redirect(url_for('admin.index'))

# Export Reports
@bp.route('/export-reports')
@login_required
@admin_required
def export_reports():
    """Export reports and data"""
    try:
        from datetime import datetime, timedelta
        from app.models.document import Document

        # Get summary statistics for export options
        total_users = User.query.count()
        total_documents = Document.query.count()
        total_categories = Category.query.count()
        total_activities = AuditLog.query.count()

        # Recent data counts
        seven_days_ago = datetime.now() - timedelta(days=7)
        recent_activities = AuditLog.query.filter(AuditLog.created_at >= seven_days_ago).count()
        recent_documents = Document.query.filter(Document.created_at >= seven_days_ago).count()

        export_options = {
            'users': {
                'total': total_users,
                'description': 'تصدير بيانات جميع المستخدمين'
            },
            'documents': {
                'total': total_documents,
                'description': 'تصدير بيانات جميع الوثائق'
            },
            'categories': {
                'total': total_categories,
                'description': 'تصدير بيانات جميع الفئات'
            },
            'activities': {
                'total': total_activities,
                'recent': recent_activities,
                'description': 'تصدير سجل الأنشطة'
            },
            'recent_documents': {
                'total': recent_documents,
                'description': 'تصدير الوثائق المضافة خلال الأسبوع الماضي'
            }
        }

        return render_template('admin/export_reports.html',
                             export_options=export_options)

    except Exception as e:
        flash(f'خطأ في تحميل صفحة التصدير: {str(e)}', 'error')
        return redirect(url_for('admin.index'))

# Export endpoints
@bp.route('/export/<export_type>')
@bp.route('/export/<export_type>/<format_type>')
@login_required
@admin_required
def export_data(export_type, format_type='csv'):
    """Export data in various formats (CSV or Excel)"""
    try:
        from datetime import datetime, timedelta
        from app.models.document import Document

        if format_type.lower() == 'excel':
            return export_to_excel(export_type)
        else:
            return export_to_csv(export_type)

    except Exception as e:
        flash(f'خطأ في تصدير البيانات: {str(e)}', 'error')
        return redirect(url_for('admin.export_reports'))

def export_to_csv(export_type):
    """Export data to CSV format"""
    try:
        import csv
        import io
        from flask import make_response
        from datetime import datetime, timedelta
        from app.models.document import Document

        # Create CSV output
        output = io.StringIO()

        if export_type == 'users':
            writer = csv.writer(output)
            writer.writerow(['ID', 'اسم المستخدم', 'البريد الإلكتروني', 'الاسم الكامل', 'الدور', 'نشط', 'تاريخ الإنشاء'])

            users = User.query.all()
            for user in users:
                writer.writerow([
                    user.id,
                    user.username,
                    user.email,
                    user.get_full_name(),
                    user.role.description if user.role else 'غير محدد',
                    'نعم' if user.is_active else 'لا',
                    user.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
            filename = f'users_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

        elif export_type == 'documents':
            writer = csv.writer(output)
            writer.writerow(['ID', 'العنوان', 'الوصف', 'نوع الملف', 'الحجم', 'الفئة', 'المؤلف', 'تاريخ الإنشاء'])

            documents = Document.query.all()
            for doc in documents:
                writer.writerow([
                    doc.id,
                    doc.title,
                    doc.description,
                    doc.file_type,
                    doc.file_size,
                    doc.category.name if doc.category else 'غير محدد',
                    doc.author.username if doc.author else 'غير محدد',
                    doc.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
            filename = f'documents_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

        elif export_type == 'categories':
            writer = csv.writer(output)
            writer.writerow(['ID', 'الاسم', 'الاسم بالعربية', 'الوصف', 'الفئة الأم', 'عدد الوثائق', 'تاريخ الإنشاء'])

            categories = Category.query.all()
            for cat in categories:
                document_count = Document.query.filter_by(category_id=cat.id).count()
                writer.writerow([
                    cat.id,
                    cat.name,
                    cat.name_ar,
                    cat.description,
                    cat.parent.name if cat.parent else 'فئة رئيسية',
                    document_count,
                    cat.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
            filename = f'categories_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

        elif export_type == 'activities':
            writer = csv.writer(output)
            writer.writerow(['ID', 'المستخدم', 'الإجراء', 'التفاصيل', 'نجح', 'عنوان IP', 'التاريخ'])

            activities = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(1000).all()
            for activity in activities:
                writer.writerow([
                    activity.id,
                    activity.username,
                    translate_action(activity.action),
                    activity.details or '',
                    'نعم' if activity.success else 'لا',
                    activity.ip_address or '',
                    activity.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
            filename = f'activities_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

        elif export_type == 'recent_documents':
            seven_days_ago = datetime.now() - timedelta(days=7)
            writer = csv.writer(output)
            writer.writerow(['ID', 'العنوان', 'الوصف', 'نوع الملف', 'الحجم', 'الفئة', 'المؤلف', 'تاريخ الإنشاء'])

            documents = Document.query.filter(Document.created_at >= seven_days_ago).all()
            for doc in documents:
                writer.writerow([
                    doc.id,
                    doc.title,
                    doc.description,
                    doc.file_type,
                    doc.file_size,
                    doc.category.name if doc.category else 'غير محدد',
                    doc.author.username if doc.author else 'غير محدد',
                    doc.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
            filename = f'recent_documents_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

        else:
            flash('نوع التصدير غير صحيح', 'error')
            return redirect(url_for('admin.export_reports'))

        # Create response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Add BOM for proper Arabic display in Excel
        response.data = '\ufeff' + response.data.decode('utf-8')

        return response

    except Exception as e:
        flash(f'خطأ في تصدير البيانات: {str(e)}', 'error')
        return redirect(url_for('admin.export_reports'))

def export_to_excel(export_type):
    """Export data to Excel format"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from flask import make_response
        from datetime import datetime, timedelta
        from app.models.document import Document
        import io

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        if export_type == 'users':
            ws.title = "بيانات المستخدمين"
            headers = ['ID', 'اسم المستخدم', 'البريد الإلكتروني', 'الاسم الكامل', 'الدور', 'نشط', 'تاريخ الإنشاء']
            ws.append(headers)

            users = User.query.all()
            for user in users:
                ws.append([
                    user.id,
                    user.username,
                    user.email,
                    user.get_full_name(),
                    user.role.description if user.role else 'غير محدد',
                    'نعم' if user.is_active else 'لا',
                    user.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
            filename = f'users_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        elif export_type == 'documents':
            ws.title = "بيانات الوثائق"
            headers = ['ID', 'العنوان', 'الوصف', 'نوع الملف', 'الحجم (بايت)', 'الفئة', 'المؤلف', 'تاريخ الإنشاء']
            ws.append(headers)

            documents = Document.query.all()
            for doc in documents:
                ws.append([
                    doc.id,
                    doc.title,
                    doc.description,
                    doc.file_type,
                    doc.file_size,
                    doc.category.name if doc.category else 'غير محدد',
                    doc.author.username if doc.author else 'غير محدد',
                    doc.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
            filename = f'documents_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        elif export_type == 'categories':
            ws.title = "بيانات الفئات"
            headers = ['ID', 'الاسم', 'الاسم بالعربية', 'الوصف', 'الفئة الأم', 'المستوى', 'عدد الوثائق', 'تاريخ الإنشاء']
            ws.append(headers)

            categories = Category.query.all()
            for cat in categories:
                document_count = Document.query.filter_by(category_id=cat.id).count()
                level = get_category_level(cat)
                ws.append([
                    cat.id,
                    cat.name,
                    cat.name_ar,
                    cat.description,
                    cat.parent.name if cat.parent else 'فئة رئيسية',
                    level,
                    document_count,
                    cat.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
            filename = f'categories_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        elif export_type == 'activities':
            ws.title = "سجل الأنشطة"
            headers = ['ID', 'المستخدم', 'الإجراء', 'التفاصيل', 'نجح', 'عنوان IP', 'التاريخ']
            ws.append(headers)

            activities = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(1000).all()
            for activity in activities:
                ws.append([
                    activity.id,
                    activity.username,
                    translate_action(activity.action),
                    activity.details or '',
                    'نعم' if activity.success else 'لا',
                    activity.ip_address or '',
                    activity.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
            filename = f'activities_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        elif export_type == 'recent_documents':
            seven_days_ago = datetime.now() - timedelta(days=7)
            ws.title = "الوثائق الحديثة"
            headers = ['ID', 'العنوان', 'الوصف', 'نوع الملف', 'الحجم (بايت)', 'الفئة', 'المؤلف', 'تاريخ الإنشاء']
            ws.append(headers)

            documents = Document.query.filter(Document.created_at >= seven_days_ago).all()
            for doc in documents:
                ws.append([
                    doc.id,
                    doc.title,
                    doc.description,
                    doc.file_type,
                    doc.file_size,
                    doc.category.name if doc.category else 'غير محدد',
                    doc.author.username if doc.author else 'غير محدد',
                    doc.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
            filename = f'recent_documents_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        else:
            flash('نوع التصدير غير صحيح', 'error')
            return redirect(url_for('admin.export_reports'))

        # Apply styles to header row
        for col_num, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Apply borders to all cells and auto-adjust column widths
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                if cell.row > 1:  # Data rows
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        flash(f'خطأ في تصدير ملف Excel: {str(e)}', 'error')
        return redirect(url_for('admin.export_reports'))

# API endpoints for dashboard
@bp.route('/api/statistics')
@login_required
@admin_required
def api_statistics():
    """API endpoint for dashboard statistics"""
    try:
        user_count = User.query.count()
        active_users = User.query.filter_by(is_active=True).count()
        category_count = Category.query.count()

        # Get document count if Document model exists
        try:
            from app.models.document import Document
            document_count = Document.query.count()
        except ImportError:
            document_count = 0

        # Get recent activities count (today)
        from datetime import datetime, timedelta
        today = datetime.now().date()
        recent_activities = AuditLog.query.filter(
            AuditLog.created_at >= today
        ).count()

        return json.dumps({
            'success': True,
            'stats': {
                'user_count': user_count,
                'active_users': active_users,
                'category_count': category_count,
                'document_count': document_count,
                'recent_activities': recent_activities
            }
        })
    except Exception as e:
        return json.dumps({'success': False, 'error': str(e)})

@bp.route('/api/logs/<int:log_id>')
@login_required
@admin_required
def api_log_details(log_id):
    """API endpoint for log details"""
    from flask import Response
    try:
        log = AuditLog.query.get_or_404(log_id)
        response_data = json.dumps({
            'success': True,
            'log': log.to_dict()
        }, ensure_ascii=False)
        return Response(response_data, content_type='application/json; charset=utf-8')
    except Exception as e:
        response_data = json.dumps({'success': False, 'error': str(e)}, ensure_ascii=False)
        return Response(response_data, content_type='application/json; charset=utf-8')

@bp.route('/api/system-status')
@login_required
@admin_required
def api_system_status():
    """API endpoint for system status"""
    try:
        # Check database connection
        db.session.execute('SELECT 1')

        # Check if we can write to database
        test_log = AuditLog.query.first()

        return json.dumps({
            'success': True,
            'status': 'online',
            'database': 'connected',
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return json.dumps({
            'success': False,
            'status': 'offline',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        })

# Additional admin routes
@bp.route('/backup')
@login_required
@super_admin_required
def backup():
    """System backup page"""
    return render_template('admin/backup.html')

@bp.route('/system-info')
@login_required
@super_admin_required
def system_info():
    """System information page"""
    import sys
    import platform
    from flask import __version__ as flask_version

    system_info = {
        'python_version': sys.version,
        'platform': platform.platform(),
        'flask_version': flask_version,
        'database_url': current_app.config.get('SQLALCHEMY_DATABASE_URI', 'Not configured'),
        'debug_mode': current_app.debug,
        'secret_key_set': bool(current_app.secret_key)
    }

    return render_template('admin/system_info.html', system_info=system_info)

